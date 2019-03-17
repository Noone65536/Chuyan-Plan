from snownlp import SnowNLP
from openpyxl import load_workbook
from openpyxl import Workbook
open = load_workbook('C:/Users/Edwar/Desktop/phone2.xlsx')
get_sheet = open.worksheets[0]
rows = get_sheet.rows

for i in range(1, 200):
        print(get_sheet.cell(row=i, column=1).value)
        s1 = SnowNLP(get_sheet.cell(row=i, column=1).value)
        print(s1.sentiments)
        rows = [s1.sentiments]
        get_sheet.append(rows)

open.save('C:/Users/Edwar/Desktop/phone2.xlsx')
